from selenium import webdriver
from selenium.webdriver.common.by import By
import urllib.request

driver = webdriver.Chrome()

# 5月29日時点のランキング(更新する必要あり)
driver.get("https://www.oricon.co.jp/music/rankinglab/dis/2023-05-29/")

detail_music = driver.find_elements(By.CLASS_NAME, "media-information")

names = []
artists = []
startdays = []

for music in detail_music:

    # 名前を取得
    music_name = music.find_element(By.CLASS_NAME, "media-title").text
    names.append(music_name)

    # 価格を取得
    music_artist = music.find_element(By.CLASS_NAME, "media-artist").text
    artists.append(music_artist)

    # 内容
    music_addinformation = music.find_element(By.CLASS_NAME, "add-information")
    music_startday = music_addinformation.find_element(By.CLASS_NAME, "media-release").text
    startdays.append(music_startday)

# excelファイルに書き込む

tmp_num = len(detail_music)
 
##以下、エクセル出力に関する部分
import openpyxl
 
wb = openpyxl.Workbook() #エクセルファイルを新規作成
sheet = wb.active
sheet.title = 'blog_title' 
 
sheet["A1"].value = '曲名'
sheet["B1"].value = '歌手名'
sheet["C1"].value = '配信開始日'
 
for i in range(1, tmp_num):
    sheet.cell(column=1, row=i+1, value=names[i-1])
    sheet.cell(column=2, row=i+1, value=artists[i-1])
    sheet.cell(column=3, row=i+1, value=startdays[i-1])

wb.save('music_excel.xlsx')
wb.close()